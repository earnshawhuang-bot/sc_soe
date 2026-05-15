"""
Excel Writer - Professional consulting-style output with full formatting.
"""
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import CellIsRule, Rule
from openpyxl.styles.differential import DifferentialStyle
from datetime import datetime


# ── Color palette (McKinsey-inspired) ──────────────────────────────────────
C_NAVY       = "1F3864"   # Header background
C_DARK_BLUE  = "2E4D8A"   # Sub-header
C_MID_BLUE   = "4472C4"   # Accent
C_LIGHT_BLUE = "D6E4F7"   # Alternate row
C_WHITE      = "FFFFFF"
C_OFF_WHITE  = "F8F9FC"
C_LIGHT_GREY = "F2F2F2"
C_BORDER     = "BFC9D6"

C_GREEN      = "D6F0E0"   # Risk tier fills
C_GREEN_FG   = "1A7A3F"
C_YELLOW     = "FFF8D6"
C_YELLOW_FG  = "8A6800"
C_ORANGE     = "FDEBD0"
C_ORANGE_FG  = "B85C00"
C_RED        = "FADADD"
C_RED_FG     = "9B1C1C"
C_CRITICAL   = "E8B4B8"
C_CRITICAL_FG= "6B0000"


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, size=10, color=C_NAVY, name="Calibri"):
    return Font(bold=bold, size=size, color=color, name=name)

def _border_thin():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def _border_bottom():
    return Border(bottom=Side(style="medium", color=C_BORDER))

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


DATE_FIELD_NAMES = {
    "planned_end_date",
    "available_date",
    "loading_date",
    "lp_earliest_valid_loading_date",
}


def _apply_date_format(cell, value):
    """Show plain dates as dates, but preserve minute-level time if present."""
    if not isinstance(value, (pd.Timestamp, datetime)):
        return
    ts = pd.Timestamp(value)
    if pd.isna(ts):
        return
    if ts.hour or ts.minute or ts.second or ts.microsecond:
        cell.number_format = "yyyy-mm-dd hh:mm"
    else:
        cell.number_format = "yyyy-mm-dd"

RISK_STYLE = {
    "Green":    (_fill(C_GREEN),    _font(bold=True, color=C_GREEN_FG)),
    "Yellow":   (_fill(C_YELLOW),   _font(bold=True, color=C_YELLOW_FG)),
    "Orange":   (_fill(C_ORANGE),   _font(bold=True, color=C_ORANGE_FG)),
    "Red":      (_fill(C_RED),      _font(bold=True, color=C_RED_FG)),
    "Critical": (_fill(C_CRITICAL), _font(bold=True, color=C_CRITICAL_FG)),
}


def write_excel(
    master: pd.DataFrame,
    output_dir: str,
    month: str,
    target_mt: float,
    data_date: str = "",
    sc_audits: dict = None,
    lp_outputs: dict = None,
    version_suffix: str = "",
):
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    suffix = f"-{version_suffix}" if version_suffix else ""
    output_path = Path(output_dir) / f"SOE_Tracking_{month}{suffix}.xlsx"

    wb = Workbook()
    wb.remove(wb.active)
    master.attrs["data_date"] = data_date

    _sheet_summary(wb, master, month, target_mt, lp_outputs or {})
    _sheet_master(wb, master)
    _sheet_gap_detail(wb, master)
    _sheet_action(wb, master, lp_outputs or {})
    _sheet_overlap_audit(wb, master)
    _sheet_sc_audits(wb, sc_audits or {})
    _sheet_loading_plan_outputs(wb, lp_outputs or {})

    wb.save(str(output_path))
    return str(output_path)


# ── Sheet 1: Executive Summary ─────────────────────────────────────────────
def _sheet_summary(wb, master, month, target_mt, lp_outputs=None):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    master = master.copy()
    master["red_critical_mt"] = master.apply(
        lambda row: row["sc_vol_mt"] if row["risk_tier"] in ["Red", "Critical"] else 0,
        axis=1,
    )

    # ── Title banner ──
    ws.merge_cells("A1:J1")
    ws.row_dimensions[1].height = 36
    c = ws["A1"]
    c.value = f"S&OE ORDER FULFILLMENT TRACKING  │  {month}"
    c.font = Font(bold=True, size=16, color=C_WHITE, name="Calibri")
    c.fill = _fill(C_NAVY)
    c.alignment = _align("center")

    ws.merge_cells("A2:J2")
    c = ws["A2"]
    c.value = f"Data as of {master.attrs.get('data_date', '')}   │   Target: {target_mt:,.0f} MT/month   │   Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    c.font = Font(size=9, color="AABBCC", name="Calibri")
    c.fill = _fill(C_NAVY)
    c.alignment = _align("center")
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 8  # spacer

    # ── KPI cards (row 4-8) ──
    _kpi_header(ws, row=4)
    kpis = [
        ("BASELINE ORDERS", f"{master['sc_vol_mt'].sum():,.0f} MT", "Adjusted baseline", C_DARK_BLUE),
        ("SHIPPED",         f"{master['allocated_shipped_mt'].sum():,.0f} MT",
                            "Allocated shipped", C_GREEN_FG),
        ("IN STOCK (FG)",   f"{master['allocated_fg_mt'].sum():,.0f} MT",
                            "Allocated FG", "2E7D32"),
        ("IN PRODUCTION",   f"{master['allocated_wip_mt'].sum():,.0f} MT",
                            "Allocated WIP", C_YELLOW_FG),
        ("SCHEDULING",      f"{(master['allocated_unsched_mt']+master['allocated_no_plan_mt']).sum():,.0f} MT",
                            "Unscheduled + no plan", C_RED_FG),
    ]
    col_starts = [1, 3, 5, 7, 9]
    for (label, val, sub, color), col in zip(kpis, col_starts):
        _kpi_card(ws, row=5, col=col, label=label, value=val, sub=sub, accent=color)

    ws.row_dimensions[10].height = 8  # spacer

    # ── Risk Distribution table (row 11+) ──
    _section_header(ws, row=11, col=1, title="RISK DISTRIBUTION", span=4)
    risk_headers = ["Risk Tier", "Volume (MT)", "% of Baseline"]
    _table_header(ws, row=12, col=1, headers=risk_headers, widths=[18, 8, 16, 16])
    baseline_mt = master["sc_vol_mt"].sum() or 1
    risk_order = ["Green", "Yellow", "Orange", "Red", "Critical"]
    for i, tier in enumerate(risk_order):
        sub = master[master["risk_tier"] == tier]
        vol = sub["sc_vol_mt"].sum()
        row_data = [tier, vol, f"{vol/baseline_mt*100:.0f}%"]
        r = 13 + i
        fill_c, font_c = RISK_STYLE.get(tier, (_fill(C_OFF_WHITE), _font()))
        for j, val in enumerate(row_data):
            c = ws.cell(row=r, column=1 + j, value=val)
            c.fill = fill_c if j == 0 else _fill(C_OFF_WHITE if i % 2 == 0 else C_LIGHT_GREY)
            c.font = font_c if j == 0 else _font()
            c.alignment = _align("center" if j > 0 else "left")
            c.border = _border_thin()
        ws.row_dimensions[r].height = 20

    # ── Plant Breakdown table ──
    _section_header(ws, row=11, col=6, title="BY PLANT", span=5)
    plant_headers = ["Plant", "Baseline MT", "Shipped", "FG", "WIP", "Unscheduled", "No Plan", "Red+Critical MT"]
    _table_header(ws, row=12, col=6, headers=plant_headers, widths=[10, 14, 12, 10, 10, 14, 12, 14])
    plants = master.groupby("plant").agg(
        baseline_mt=("sc_vol_mt", "sum"),
        shipped_mt=("shipped_mt", "sum"),
        fg_mt=("fg_mt", "sum"),
        wip_mt=("wip_mt", "sum"),
        unsched_mt=("unsched_mt", "sum"),
        no_plan_mt=("no_plan_mt", "sum"),
        risk_red_mt=("red_critical_mt", "sum"),
    ).reset_index()
    for i, (_, row) in enumerate(plants.iterrows()):
        r = 13 + i
        vals = [row["plant"], row["baseline_mt"],
                row["shipped_mt"], row["fg_mt"],
                row["wip_mt"], row["unsched_mt"],
                row["no_plan_mt"],
                row["risk_red_mt"]]
        bg = C_OFF_WHITE if i % 2 == 0 else C_LIGHT_GREY
        for j, val in enumerate(vals):
            c = ws.cell(row=r, column=6 + j, value=val)
            c.fill = _fill(bg)
            c.font = _font(bold=(j == 0))
            c.alignment = _align("center" if j > 0 else "left")
            c.border = _border_thin()
        ws.row_dimensions[r].height = 20

    # ── Risk Tier Logic ──
    logic_start = 20
    _section_header(ws, row=logic_start, col=6, title="RISK TIER LOGIC", span=6)
    _table_header(ws, row=logic_start+1, col=6, headers=["Risk Tier", "Trigger"], widths=[14, 72])
    logic_rows = [
        ("Green", "Fully shipped, or in production with gap > 2 days"),
        ("Yellow", "In stock / partial shipped, or in production with gap 0-2 days"),
        ("Orange", "Allocated WIP exists but no loading plan date"),
        ("Red", "Allocated unscheduled work order exists, or production gap < 0"),
        ("Critical", "Allocated no-plan quantity remains after shipped, FG, WIP, and unscheduled"),
    ]
    for i, (tier, trigger) in enumerate(logic_rows):
        r = logic_start + 2 + i
        fill_c, font_c = RISK_STYLE.get(tier, (_fill(C_OFF_WHITE), _font()))
        for j, val in enumerate([tier, trigger]):
            c = ws.cell(row=r, column=6+j, value=val)
            c.fill = fill_c if j == 0 else _fill(C_OFF_WHITE if i % 2 == 0 else C_LIGHT_GREY)
            c.font = font_c if j == 0 else _font(size=9)
            c.alignment = _align("left", wrap=(j == 1))
            c.border = _border_thin()
        ws.row_dimensions[r].height = 22

    # Loading Plan risk matrix: LP is a shipping-arrangement axis, not a
    # standalone execution bucket in the Summary narrative.
    if lp_outputs:
        _section_header(ws, row=11, col=14, title="SHIPPED DATA CLOSURE (MT)", span=4)
        _table_header(
            ws,
            row=12,
            col=14,
            headers=["Supply Status", "Total MT", "LP Closed / Matched", "LP Missing or Unclear"],
            widths=[22, 16, 20, 22],
        )
        for i, row_data in enumerate(_build_shipped_closure_matrix(lp_outputs)):
            r = 13 + i
            bg = C_OFF_WHITE if i % 2 == 0 else C_LIGHT_GREY
            for j, val in enumerate(row_data):
                c = ws.cell(row=r, column=14+j, value=val)
                c.fill = _fill(bg)
                c.font = _font(bold=(j == 0))
                c.alignment = _align("center" if j > 0 else "left", wrap=True)
                c.border = _border_thin()
            ws.row_dimensions[r].height = 28

        _section_header(ws, row=16, col=14, title="OPEN FULFILLMENT RISK MATRIX (MT)", span=6)
        _table_header(
            ws,
            row=17,
            col=14,
            headers=[
                "Supply Status",
                "Supply Status Total",
                "Past Due LP",
                "Future Valid LP",
                "LP Date Unconfirmed",
                "No Current LP",
            ],
            widths=[22, 18, 16, 18, 20, 18],
        )
        for i, row_data in enumerate(_build_lp_risk_matrix(master, lp_outputs)):
            r = 18 + i
            bg = C_OFF_WHITE if i % 2 == 0 else C_LIGHT_GREY
            for j, val in enumerate(row_data):
                c = ws.cell(row=r, column=14+j, value=val)
                c.fill = _fill(bg)
                c.font = _font(bold=(j == 0))
                c.alignment = _align("center" if j > 0 else "left", wrap=True)
                c.border = _border_thin()
            ws.row_dimensions[r].height = 30

    # ── Order Type Breakdown ──
    start_row = 29
    _section_header(ws, row=start_row, col=1, title="ORDER TYPE BREAKDOWN", span=9)
    ot_headers = ["Order Type", "Volume (MT)", "Shipped", "FG", "WIP", "Unscheduled", "No Plan"]
    _table_header(ws, row=start_row+1, col=1, headers=ot_headers, widths=[28,14,12,10,10,14,12])
    for i, row in enumerate(_build_order_type_breakdown(master)):
        vals = [
            row["order_type"], row["volume_mt"],
            row["shipped_mt"], row["fg_mt"],
            row["wip_mt"], row["unsched_mt"],
            row["no_plan_mt"],
        ]
        r = start_row + 2 + i
        bg = C_OFF_WHITE if i % 2 == 0 else C_LIGHT_GREY
        for j, val in enumerate(vals):
            c = ws.cell(row=r, column=1+j, value=val)
            c.fill = _fill(bg)
            c.font = _font(bold=(j==0))
            c.alignment = _align("center" if j > 0 else "left")
            c.border = _border_thin()
        ws.row_dimensions[r].height = 20

    # ── Plant × Region (Cluster) Breakdown ──
    # start_row + 1 (section) + 1 (headers) + 3 (data) + 1 (spacer) = start_row + 6
    pc_start = start_row + 6
    ws.row_dimensions[pc_start - 1].height = 10  # spacer row before this section
    _section_header(ws, row=pc_start, col=1, title="BY PLANT × REGION (CLUSTER)", span=9)
    pc_headers = ["Plant", "Cluster", "Baseline MT", "Shipped", "FG", "WIP", "Unscheduled", "No Plan", "Red+Critical MT", "% Fulfilled"]
    _table_header(ws, row=pc_start+1, col=1, headers=pc_headers, widths=[10,12,14,12,10,10,14,12,14,12])

    plant_cluster_detail = master.groupby(["plant", "cluster"]).agg(
        baseline_mt=("sc_vol_mt", "sum"),
        shipped_mt=("shipped_mt", "sum"),
        fg_mt=("fg_mt", "sum"),
        wip_mt=("wip_mt", "sum"),
        unsched_mt=("unsched_mt", "sum"),
        no_plan_mt=("no_plan_mt", "sum"),
        risk_red_mt=("red_critical_mt", "sum"),
    ).reset_index().sort_values(["plant", "cluster"])
    plant_totals = master.groupby("plant").agg(
        baseline_mt=("sc_vol_mt", "sum"),
        shipped_mt=("shipped_mt", "sum"),
        fg_mt=("fg_mt", "sum"),
        wip_mt=("wip_mt", "sum"),
        unsched_mt=("unsched_mt", "sum"),
        no_plan_mt=("no_plan_mt", "sum"),
        risk_red_mt=("red_critical_mt", "sum"),
    ).reset_index()
    plant_totals["cluster"] = "TOTAL"
    plant_cluster = pd.concat([plant_totals, plant_cluster_detail], ignore_index=True)
    plant_cluster["_cluster_sort"] = plant_cluster["cluster"].apply(lambda x: "" if x == "TOTAL" else str(x))
    plant_cluster = plant_cluster.sort_values(["plant", "_cluster_sort"]).drop(columns=["_cluster_sort"])

    last_plant = None
    for i, (_, row) in enumerate(plant_cluster.iterrows()):
        r = pc_start + 2 + i
        fulfilled = row["shipped_mt"] + row["fg_mt"]
        pct = f"{fulfilled / row['baseline_mt'] * 100:.0f}%" if row["baseline_mt"] > 0 else "—"
        plant_label = row["plant"] if row["plant"] != last_plant else ""
        last_plant = row["plant"]
        vals = [plant_label, row["cluster"],
                row["baseline_mt"], row["shipped_mt"],
                row["fg_mt"], row["wip_mt"],
                row["unsched_mt"],
                row["no_plan_mt"],
                row["risk_red_mt"], pct]
        bg = C_LIGHT_BLUE if row["cluster"] == "TOTAL" else (C_OFF_WHITE if i % 2 == 0 else C_LIGHT_GREY)
        for j, val in enumerate(vals):
            c = ws.cell(row=r, column=1+j, value=val)
            c.fill = _fill(bg)
            c.font = _font(bold=(row["cluster"] == "TOTAL" or (j <= 1 and bool(plant_label))))
            c.alignment = _align("center" if j > 1 else "left")
            c.border = _border_thin()
        ws.row_dimensions[r].height = 20

    # Column widths for summary sheet
    for col in range(1, 13):
        ws.column_dimensions[get_column_letter(col)].width = 14
    _format_summary_numbers(ws)


def _build_shipped_closure_matrix(lp_outputs):
    seg = lp_outputs.get("risk_matrix_detail", pd.DataFrame())
    if seg is None or seg.empty:
        return [["Shipped", 0, 0, 0]]
    shipped = seg[seg["supply_status"] == "Shipped"].copy()
    total = pd.to_numeric(shipped.get("risk_mt", pd.Series(dtype=float)), errors="coerce").fillna(0).sum()
    closed = pd.to_numeric(
        shipped.loc[shipped["lp_coverage_status"] == "Shipped LP Closed", "risk_mt"],
        errors="coerce",
    ).fillna(0).sum()
    unclear = total - closed
    return [["Shipped", float(total) if total else 0, float(closed) if closed else 0, float(unclear) if unclear else 0]]


def _build_lp_risk_matrix(master, lp_outputs):
    seg = lp_outputs.get("risk_matrix_detail", pd.DataFrame())
    if seg is None:
        seg = pd.DataFrame()

    supply_order = ["FG", "WIP Scheduled", "WIP Unscheduled", "No Supply Signal"]
    lp_order = ["Past Due LP", "Future Valid LP", "LP Date Unconfirmed", "No Current LP"]
    rows = []
    for supply in supply_order:
        values = []
        for lp_status in lp_order:
            if seg.empty:
                mt = 0
            else:
                sub = seg[(seg["supply_status"] == supply) & (seg["lp_coverage_status"] == lp_status)]
                mt = pd.to_numeric(sub.get("risk_mt", pd.Series(dtype=float)), errors="coerce").fillna(0).sum()
            values.append(float(mt) if mt else 0)
        row = [supply, sum(values)] + values
        rows.append(row)
    if rows:
        totals = ["Open Total"]
        for idx in range(1, len(rows[0])):
            totals.append(sum(row[idx] for row in rows))
        rows.append(totals)
    return rows


def _format_summary_numbers(ws):
    """Show Summary numeric values as thousands-formatted integers."""
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'


def _build_order_type_breakdown(master):
    """Allocate SO-level waterfall quantities to row-level order types by volume share."""
    order_type_fields = [
        ("Carry Over Stock", "carry_over_stock_mt"),
        ("Carry Over Unproduced", "carry_over_unproduced_mt"),
        ("Fresh Order This Month", "fresh_this_month_mt"),
    ]
    allocated_fields = {
        "shipped_mt": "allocated_shipped_mt",
        "fg_mt": "allocated_fg_mt",
        "wip_mt": "allocated_wip_mt",
        "unsched_mt": "allocated_unsched_mt",
        "no_plan_mt": "allocated_no_plan_mt",
    }
    rows = []
    for order_type, vol_field in order_type_fields:
        if vol_field not in master.columns:
            rows.append({
                "order_type": order_type, "so_count": 0, "volume_mt": 0,
                "shipped_mt": 0, "fg_mt": 0, "wip_mt": 0, "unsched_mt": 0, "no_plan_mt": 0,
            })
            continue
        type_volume = master[vol_field].fillna(0)
        share = (type_volume / master["sc_vol_mt"].replace(0, pd.NA)).fillna(0)
        row = {
            "order_type": order_type,
            "so_count": int((type_volume > 0).sum()),
            "volume_mt": type_volume.sum(),
        }
        for output_field, source_field in allocated_fields.items():
            row[output_field] = (master[source_field] * share).sum()
        rows.append(row)
    return rows


# ── Sheet 2: SO Master ─────────────────────────────────────────────────────
def _sheet_master(wb, master):
    ws = wb.create_sheet("SO Master")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    _banner(ws, "SO MASTER — ALL BASELINE ORDERS", span=30)

    col_defs = [
        ("SO Number",        "so",               12),
        ("Plant",            "plant",              8),
        ("Cluster",          "cluster",           10),
        ("Order Type",       "order_type",        22),
        ("Adjusted SC Vol (MT)", "sc_vol_mt",         16),
        ("Raw SC Vol (MT)",      "raw_sc_vol_mt",     14),
        ("Allocated SC Prior Shipped", "allocated_sc_prior_shipped_mt", 24),
        ("Allocated GI Shipped", "allocated_gi_shipped_mt", 20),
        ("Allocated Shipped",    "allocated_shipped_mt", 16),
        ("Allocated FG",         "allocated_fg_mt",      14),
        ("Allocated WIP",        "allocated_wip_mt",     14),
        ("Allocated Unscheduled","allocated_unsched_mt", 18),
        ("Allocated No Plan",    "allocated_no_plan_mt", 16),
        ("Raw SC Prior Delivery","raw_sc_prior_delivery_mt", 22),
        ("Raw Shipped",          "raw_shipped_mt",       12),
        ("Raw FG",               "raw_fg_mt",            12),
        ("Raw WIP",              "raw_wip_mt",           12),
        ("Raw Unscheduled",      "raw_unsched_mt",       16),
        ("CO Stock Adj",         "carry_over_stock_mt",  14),
        ("CO Unprod Adj",        "carry_over_unproduced_mt", 15),
        ("Fresh Adj",            "fresh_this_month_mt",  12),
        ("Status",           "status",            22),
        ("Risk Tier",        "risk_tier",         12),
        ("Planned End Date", "planned_end_date",  16),
        ("Available Date",   "available_date",    16),
        ("Loading Date",     "loading_date",      14),
        ("Gap (days)",       "gap_days",          12),
        ("Machines",         "machines",          20),
        ("LP Source",        "lp_source",         14),
    ]
    headers = [h for h, _, _ in col_defs]
    widths  = [w for _, _, w in col_defs]
    fields  = [f for _, f, _ in col_defs]

    _table_header(ws, row=2, col=1, headers=headers, widths=widths)

    # Sort: Critical → Red → Orange → Yellow → Green
    risk_priority = {"Critical": 0, "Red": 1, "Orange": 2, "Yellow": 3, "Green": 4}
    sorted_df = master.copy()
    sorted_df["_rp"] = sorted_df["risk_tier"].map(risk_priority).fillna(5)
    sorted_df = sorted_df.sort_values(["_rp", "gap_days"], ascending=[True, True])

    for i, (_, row) in enumerate(sorted_df.iterrows()):
        r = 3 + i
        bg = C_OFF_WHITE if i % 2 == 0 else C_LIGHT_GREY
        for j, field in enumerate(fields):
            val = row.get(field)
            if pd.isna(val) if not isinstance(val, str) else False:
                val = ""
            if isinstance(val, float) and val == int(val) and field.endswith("_mt"):
                val = round(val, 1)
            c = ws.cell(row=r, column=1+j, value=val)
            c.fill = _fill(bg)
            c.font = _font()
            c.alignment = _align("center" if j > 1 else "left")
            c.border = _border_thin()
            # Risk tier column coloring
            if field == "risk_tier" and val in RISK_STYLE:
                c.fill, c.font = RISK_STYLE[val]
            # Gap coloring
            if field == "gap_days" and isinstance(val, (int, float)):
                if val < 0:
                    c.fill = _fill(C_RED); c.font = _font(bold=True, color=C_RED_FG)
                elif val <= 2:
                    c.fill = _fill(C_YELLOW); c.font = _font(bold=True, color=C_YELLOW_FG)
            if field in DATE_FIELD_NAMES:
                _apply_date_format(c, val)
        ws.row_dimensions[r].height = 18

    ws.auto_filter.ref = f"A2:{get_column_letter(len(col_defs))}2"


# ── Sheet 3: Gap Detail ────────────────────────────────────────────────────
def _sheet_gap_detail(wb, master):
    ws = wb.create_sheet("Gap Analysis")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    gap_df = master[
        (master["status"] == "In Production") & master["gap_days"].notna()
    ].sort_values("gap_days").copy()

    _banner(ws, f"GAP ANALYSIS — IN PRODUCTION SOs ({len(gap_df)} orders)", span=11)

    col_defs = [
        ("SO Number",        "so",               12),
        ("Plant",            "plant",              8),
        ("Cluster",          "cluster",           10),
        ("Order Type",       "order_type",        22),
        ("Adjusted SC Vol (MT)", "sc_vol_mt",         16),
        ("Allocated WIP",        "allocated_wip_mt",  14),
        ("Planned End Date", "planned_end_date",  16),
        ("Available Date",   "available_date",    16),
        ("Loading Date",     "loading_date",      14),
        ("Gap (days)",       "gap_days",          12),
        ("Risk Tier",        "risk_tier",         12),
    ]
    headers = [h for h, _, _ in col_defs]
    widths  = [w for _, _, w in col_defs]
    fields  = [f for _, f, _ in col_defs]

    _table_header(ws, row=2, col=1, headers=headers, widths=widths)

    for i, (_, row) in enumerate(gap_df.iterrows()):
        r = 3 + i
        bg = C_OFF_WHITE if i % 2 == 0 else C_LIGHT_GREY
        for j, field in enumerate(fields):
            val = row.get(field)
            if pd.isna(val) if not isinstance(val, str) else False:
                val = ""
            c = ws.cell(row=r, column=1+j, value=val)
            c.fill = _fill(bg)
            c.font = _font()
            c.alignment = _align("center" if j > 1 else "left")
            c.border = _border_thin()
            if field == "gap_days" and isinstance(val, (int, float)):
                if val < 0:
                    c.fill = _fill(C_RED); c.font = _font(bold=True, color=C_RED_FG)
                elif val <= 2:
                    c.fill = _fill(C_YELLOW); c.font = _font(bold=True, color=C_YELLOW_FG)
                else:
                    c.fill = _fill(C_GREEN); c.font = _font(bold=True, color=C_GREEN_FG)
            if field == "risk_tier" and val in RISK_STYLE:
                c.fill, c.font = RISK_STYLE[val]
            if field in DATE_FIELD_NAMES:
                _apply_date_format(c, val)
        ws.row_dimensions[r].height = 18

    ws.auto_filter.ref = f"A2:{get_column_letter(len(col_defs))}2"


# ── Sheet 4: Action Required ───────────────────────────────────────────────
def _sheet_action(wb, master, lp_outputs=None):
    ws = wb.create_sheet("Action Required")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    risk_detail = (lp_outputs or {}).get("risk_matrix_detail", pd.DataFrame())
    if risk_detail is None:
        risk_detail = pd.DataFrame()
    if risk_detail.empty or "action_required" not in risk_detail.columns:
        action_df = pd.DataFrame()
    else:
        action_df = risk_detail[risk_detail["action_required"]].copy()
        action_df = action_df.sort_values(["risk_action", "risk_mt"], ascending=[True, False])

    _banner(ws, f"ACTION REQUIRED — {len(action_df)} RISK SEGMENTS NEED FOLLOW-UP", span=15, urgent=True)

    col_defs = [
        ("SO Number",        "so",               12),
        ("Plant",            "plant",              8),
        ("Cluster",          "cluster",           10),
        ("Order Type",       "order_type",        22),
        ("SO Total MT",      "so_total_mt",       14),
        ("Risk MT",          "risk_mt",           12),
        ("Covered MT",       "covered_mt",        12),
        ("Supply Status",    "supply_status",     18),
        ("LP Coverage Status", "lp_coverage_status", 22),
        ("Risk Action",      "risk_action",       30),
        ("Owner",            "suggested_owner",   12),
        ("Loading Date Raw", "lp_loading_date_raw_list", 22),
        ("Planned End Date", "planned_end_date",  16),
        ("Available Date",   "available_date",    16),
        ("Gap Days",         "lp_gap_days",       10),
        ("Action Note",      "action_note",       44),
    ]
    headers = [h for h, _, _ in col_defs]
    widths  = [w for _, _, w in col_defs]
    fields  = [f for _, f, _ in col_defs]

    _table_header(ws, row=2, col=1, headers=headers, widths=widths)

    if action_df.empty:
        ws.cell(row=3, column=1, value="No action-required risk segments")
        ws["A3"].font = _font(bold=True)
        return

    for i, (_, row) in enumerate(action_df.iterrows()):
        r = 3 + i
        bg = C_OFF_WHITE if i % 2 == 0 else C_LIGHT_GREY
        for j, field in enumerate(fields):
            val = row.get(field)
            if pd.isna(val) if not isinstance(val, str) else False:
                val = ""
            c = ws.cell(row=r, column=1+j, value=val)
            c.fill = _fill(bg)
            c.font = _font()
            c.alignment = _align("center" if j in [5, 6, 7, 14] else "left", wrap=(field == "action_note"))
            c.border = _border_thin()
            if isinstance(val, (int, float)) and (field.endswith("_mt") or field == "risk_mt"):
                c.number_format = "#,##0.0"
            if field in DATE_FIELD_NAMES:
                _apply_date_format(c, val)
        ws.row_dimensions[r].height = 18

    ws.auto_filter.ref = f"A2:{get_column_letter(len(col_defs))}2"


def _sheet_overlap_audit(wb, master):
    ws = wb.create_sheet("Overlap Audit")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    overlap = master.copy()
    overlap["raw_coverage_mt"] = (
        overlap["raw_sc_prior_delivery_mt"]
        + overlap["raw_shipped_mt"]
        + overlap["raw_fg_mt"]
        + overlap["raw_wip_mt"]
        + overlap["raw_unsched_mt"]
    )
    overlap["raw_over_baseline_mt"] = (overlap["raw_coverage_mt"] - overlap["sc_vol_mt"]).clip(lower=0)
    overlap = overlap[overlap["raw_over_baseline_mt"] > 0.01].sort_values("raw_over_baseline_mt", ascending=False)

    _banner(ws, f"RAW SOURCE OVERLAP AUDIT — {len(overlap)} SOs", span=12)
    col_defs = [
        ("SO Number", "so", 12),
        ("Plant", "plant", 8),
        ("Cluster", "cluster", 10),
        ("Order Type", "order_type", 22),
        ("Adjusted Baseline", "sc_vol_mt", 16),
        ("Raw SC Prior Delivery", "raw_sc_prior_delivery_mt", 22),
        ("Raw Shipped", "raw_shipped_mt", 12),
        ("Raw FG", "raw_fg_mt", 12),
        ("Raw WIP", "raw_wip_mt", 12),
        ("Raw Unscheduled", "raw_unsched_mt", 16),
        ("Raw Coverage", "raw_coverage_mt", 14),
        ("Raw Over Baseline", "raw_over_baseline_mt", 18),
    ]
    _write_table(ws, overlap, col_defs, row=2)


def _sheet_sc_audits(wb, sc_audits):
    audit_specs = [
        ("SC Row Detail", "sc_row_detail"),
        ("SC Fresh Next Month", "fresh_next_month"),
        ("SC Unknown Type", "unknown_type"),
        ("SC Prior Delivery", "sc_prior_delivery"),
        ("Unmatched End Customer", "unmatched_customer"),
    ]
    for sheet_name, key in audit_specs:
        df = sc_audits.get(key)
        if df is None:
            df = pd.DataFrame()
        ws = wb.create_sheet(sheet_name[:31])
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"
        if df.empty:
            ws.cell(row=1, column=1, value="No rows")
            ws["A1"].font = _font(bold=True)
            continue
        headers = list(df.columns)
        widths = [min(max(len(str(h)) + 2, 12), 24) for h in headers]
        _table_header(ws, row=1, col=1, headers=headers, widths=widths)
        for i, (_, row) in enumerate(df.iterrows(), start=2):
            bg = C_OFF_WHITE if i % 2 == 0 else C_LIGHT_GREY
            for j, h in enumerate(headers, start=1):
                val = row[h]
                if pd.isna(val) if not isinstance(val, str) else False:
                    val = ""
                c = ws.cell(row=i, column=j, value=val)
                c.fill = _fill(bg)
                c.font = _font()
                c.border = _border_thin()
                c.alignment = _align("center" if j > 1 else "left")
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _sheet_loading_plan_outputs(wb, lp_outputs):
    """Write Loading Plan clean detail, reconciliation, readiness, and audits."""
    sheet_specs = [
        ("Loading Plan Clean Detail", "clean_detail"),
        ("SC vs LP Reconciliation", "reconciliation"),
        ("Risk Matrix Detail", "risk_matrix_detail"),
        ("LP Only Summary", "lp_not_in_baseline_summary"),
        ("LP Not In Current Baseline", "lp_not_in_baseline_detail"),
        ("Shipping Readiness", "shipping_readiness"),
        ("LP Date Exceptions", "date_exceptions"),
        ("LP Parse Exceptions", "parse_exceptions"),
        ("LP Excluded Prior Invoiced", "excluded_prior_invoiced"),
    ]
    for sheet_name, key in sheet_specs:
        df = lp_outputs.get(key)
        if df is None:
            df = pd.DataFrame()
        _write_dataframe_sheet(wb, sheet_name, df)


def _write_dataframe_sheet(wb, sheet_name, df):
    ws = wb.create_sheet(sheet_name[:31])
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    if df.empty:
        ws.cell(row=1, column=1, value="No rows")
        ws["A1"].font = _font(bold=True)
        return

    df = df.copy()
    headers = list(df.columns)
    widths = [min(max(len(str(h)) + 2, 12), 28) for h in headers]
    _table_header(ws, row=1, col=1, headers=headers, widths=widths)
    for i, (_, row) in enumerate(df.iterrows(), start=2):
        bg = C_OFF_WHITE if i % 2 == 0 else C_LIGHT_GREY
        for j, h in enumerate(headers, start=1):
            val = row[h]
            if pd.isna(val) if not isinstance(val, str) else False:
                val = ""
            c = ws.cell(row=i, column=j, value=val)
            c.fill = _fill(bg)
            c.font = _font()
            c.border = _border_thin()
            c.alignment = _align("center" if j > 1 else "left", wrap=False)
            if isinstance(val, (int, float)) and (str(h).endswith("_mt") or str(h) in ["load_mt", "source_mt"]):
                c.number_format = "#,##0.0"
            if str(h) in DATE_FIELD_NAMES:
                _apply_date_format(c, val)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _write_table(ws, df, col_defs, row=1):
    headers = [h for h, _, _ in col_defs]
    widths = [w for _, _, w in col_defs]
    fields = [f for _, f, _ in col_defs]
    _table_header(ws, row=row, col=1, headers=headers, widths=widths)
    for i, (_, record) in enumerate(df.iterrows(), start=row + 1):
        bg = C_OFF_WHITE if i % 2 == 0 else C_LIGHT_GREY
        for j, field in enumerate(fields, start=1):
            val = record.get(field)
            if pd.isna(val) if not isinstance(val, str) else False:
                val = ""
            c = ws.cell(row=i, column=j, value=val)
            c.fill = _fill(bg)
            c.font = _font()
            c.alignment = _align("center" if j > 2 else "left")
            c.border = _border_thin()
            if isinstance(val, (int, float)) and field.endswith("_mt"):
                c.number_format = "#,##0.0"
    ws.auto_filter.ref = f"A{row}:{get_column_letter(len(col_defs))}{row}"


# ── Helper functions ────────────────────────────────────────────────────────
def _banner(ws, title, span=10, urgent=False):
    ws.merge_cells(f"A1:{get_column_letter(span)}1")
    c = ws["A1"]
    c.value = title
    c.font = Font(bold=True, size=13, color=C_WHITE, name="Calibri")
    c.fill = _fill("7B0000" if urgent else C_NAVY)
    c.alignment = _align("left")
    ws.row_dimensions[1].height = 30


def _section_header(ws, row, col, title, span=4):
    end_col = get_column_letter(col + span - 1)
    ws.merge_cells(f"{get_column_letter(col)}{row}:{end_col}{row}")
    c = ws.cell(row=row, column=col, value=title)
    c.font = Font(bold=True, size=10, color=C_WHITE, name="Calibri")
    c.fill = _fill(C_DARK_BLUE)
    c.alignment = _align("left")
    ws.row_dimensions[row].height = 22


def _table_header(ws, row, col, headers, widths):
    for j, (h, w) in enumerate(zip(headers, widths)):
        c = ws.cell(row=row, column=col+j, value=h)
        c.font = Font(bold=True, size=9, color=C_WHITE, name="Calibri")
        c.fill = _fill(C_MID_BLUE)
        c.alignment = _align("center")
        c.border = _border_thin()
        ws.column_dimensions[get_column_letter(col+j)].width = w
    ws.row_dimensions[row].height = 22


def _kpi_header(ws, row):
    ws.row_dimensions[row].height = 6


def _kpi_card(ws, row, col, label, value, sub, accent):
    # Label row
    ws.merge_cells(f"{get_column_letter(col)}{row}:{get_column_letter(col+1)}{row}")
    c = ws.cell(row=row, column=col, value=label)
    c.font = Font(bold=True, size=8, color="888888", name="Calibri")
    c.fill = _fill(C_OFF_WHITE)
    c.alignment = _align("center")
    c.border = Border(top=Side(style="medium", color=accent))
    ws.row_dimensions[row].height = 18

    # Value row
    ws.merge_cells(f"{get_column_letter(col)}{row+1}:{get_column_letter(col+1)}{row+1}")
    c = ws.cell(row=row+1, column=col, value=value)
    c.font = Font(bold=True, size=18, color=accent, name="Calibri")
    c.fill = _fill(C_OFF_WHITE)
    c.alignment = _align("center")
    ws.row_dimensions[row+1].height = 30

    # Sub row
    ws.merge_cells(f"{get_column_letter(col)}{row+2}:{get_column_letter(col+1)}{row+2}")
    c = ws.cell(row=row+2, column=col, value=sub)
    c.font = Font(size=9, color="888888", name="Calibri")
    c.fill = _fill(C_OFF_WHITE)
    c.alignment = _align("center")
    ws.row_dimensions[row+2].height = 16

    # Bottom border spacer
    ws.merge_cells(f"{get_column_letter(col)}{row+3}:{get_column_letter(col+1)}{row+3}")
    c = ws.cell(row=row+3, column=col)
    c.fill = _fill(C_OFF_WHITE)
    ws.row_dimensions[row+3].height = 6
